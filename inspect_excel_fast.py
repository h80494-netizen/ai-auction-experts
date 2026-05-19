import openpyxl
import json
import os

file_path = "c:/Users/llll/Documents/두인경매/바이브코딩/data/경공매_260515.xlsx"
print(f"Reading {file_path} using read_only mode...")

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print("Sheets found:", wb.sheetnames)
    
    summary = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 4: # Read first 4 rows
                break
            data.append(list(row))
        
        if data:
            columns = data[0] if len(data) > 0 else []
            samples = [dict(zip(columns, row)) for row in data[1:]] if len(data) > 1 else []
            summary[sheet_name] = {
                "columns": columns,
                "samples": samples
            }
    
    with open("c:/Users/llll/Documents/두인경매/바이브코딩/data/auction_schema_summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
        
    print("Schema saved to data/gps_schema_summary.json")
except Exception as e:
    print("Error:", str(e))
