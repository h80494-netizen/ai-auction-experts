import sqlite3
import openpyxl

excel_path = 'data/경공매데이터_update_대항력.xlsx'
db_path = 'backend/data/map_data.db'

print("Loading Excel using openpyxl...")
wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
case_idx = headers.index('사건번호')
date_idx = headers.index('입찰일')

date_map = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    case_no = str(row[case_idx]).strip() if row[case_idx] else ''
    date_val = str(row[date_idx]).strip() if row[date_idx] else ''
    if case_no and date_val and date_val != 'nan' and date_val != 'None':
        date_map[case_no] = date_val[:10]

print(f"Found {len(date_map)} unique case numbers with dates.")

print("Updating DB...")
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

updated = 0
for case_no, sale_date in date_map.items():
    cursor.execute("UPDATE auctions SET sale_date = ? WHERE case_no = ?", (sale_date, case_no))
    updated += cursor.rowcount

conn.commit()
print(f"Updated {updated} rows in DB.")

# Verify
cursor.execute("SELECT sale_date, COUNT(*) FROM auctions GROUP BY sale_date ORDER BY COUNT(*) DESC LIMIT 10")
for r in cursor.fetchall():
    print(r)

conn.close()
