import openpyxl
import os
import datetime

file_path = "c:/Users/llll/Documents/두인경매/바이브코딩/data/경공매데이터_260515.xlsx"
print(f"Reading {file_path} using openpyxl in read_only mode...")

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active or wb[wb.sheetnames[0]]
    print(f"Sheet name: {sheet.title}")
    
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 3:
            print(f"Row {i}: {row[:15]}")
        rows.append(list(row))
        if i >= 10:
            break
            
except Exception as e:
    print("Error:", str(e))
